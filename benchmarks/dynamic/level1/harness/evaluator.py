"""Compute per-test metrics and pass/fail from an archived DMView suite run.

DMView produces no machine-readable verdicts, so this module is the benchmark's
referee: it reads the PSS/e channel (.out) files, computes waveform metrics,
and compares them against thresholds in criteria.json. Thresholds were
calibrated so that the corrupted model (10,50,10,50) fails with the symptoms
reported in the paper while the known-good gains (1,5,1,5) pass every test.
"""
import json
import math
import re
from pathlib import Path

import numpy as np

import channels
import config


# --------------------------------------------------------------------------
# criteria / binding helpers
# --------------------------------------------------------------------------

def load_criteria() -> dict:
    return json.loads(config.CRITERIA_FILE.read_text(encoding="utf-8"))


class EvalError(Exception):
    pass


def _find_out_file(results_dir: Path, test_name: str, criteria: dict) -> Path | None:
    pattern = criteria["outfile_patterns"][test_name]
    matches = [p for p in sorted(results_dir.rglob("*.out"))
               if re.search(pattern, p.name, re.IGNORECASE)]
    # DMView duplicates each test folder under a nested <PROJECT>\<PROJECT>\
    # tree; identical filenames are the same artifact - keep the shallowest.
    by_name = {}
    for p in matches:
        cur = by_name.get(p.name)
        if cur is None or len(p.parts) < len(cur.parts):
            by_name[p.name] = p
    if len(by_name) > 1:
        raise EvalError(f"{test_name}: outfile pattern {pattern!r} matched "
                        f"{sorted(by_name)}")
    return next(iter(by_name.values())) if by_name else None


def _resolve_channel(chans: dict, pattern: str, what: str) -> str:
    matches = [k for k in chans if re.search(pattern, k, re.IGNORECASE)]
    if len(matches) != 1:
        raise EvalError(f"channel pattern {pattern!r} for {what} matched "
                        f"{len(matches)} channels: {matches[:5]}")
    return matches[0]


def _signal(chans: dict, criteria: dict, kind: str, out_name: str) -> np.ndarray:
    key = _resolve_channel(chans, criteria["channel_patterns"][kind],
                           f"{kind} in {out_name}")
    return np.asarray(chans[key], dtype=float)


# --------------------------------------------------------------------------
# disturbance profiles (xlsx) -> windows
# --------------------------------------------------------------------------

def read_profile(xlsx_path: Path) -> list[tuple[float, float]]:
    """Read (time, value) breakpoints from a DMView disturbance profile."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    points = []
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 2:
            continue
        t, v = row[0], row[1]
        if isinstance(t, (int, float)) and isinstance(v, (int, float)):
            points.append((float(t), float(v)))
    wb.close()
    if len(points) < 2:
        raise EvalError(f"profile {xlsx_path.name}: found {len(points)} breakpoints")
    return points


def profile_path(test_name: str) -> Path | None:
    spec = dict(config.TESTS)[test_name]
    m = re.search(r"DATAs\\+([^']+)'", spec)
    if not m:
        return None
    return config.DMVIEW_ROOT / "DATAs" / m.group(1)


def steady_tails(points, t_end: float, settle_s: float, tail_s: float):
    """Windows over the last `tail_s` of each constant-value profile hold.

    Measuring only the tail of each hold separates a response that settles
    (legitimate transient decays before the hold ends -> tiny p-p) from a
    sustained oscillation (limit cycle persists into the tail -> large p-p).
    Ramps and holds shorter than settle_s + tail_s are skipped.
    """
    pts = sorted(points) + [(t_end, points[-1][1])]
    # merge maximal runs of equal value into holds
    holds = []
    i = 0
    while i < len(pts) - 1:
        if pts[i + 1][1] == pts[i][1]:
            j = i + 1
            while j + 1 < len(pts) and pts[j + 1][1] == pts[i][1]:
                j += 1
            holds.append((pts[i][0], pts[j][0]))
            i = j
        else:
            i += 1
    return [(b - tail_s, b) for a, b in holds if b - a >= settle_s + tail_s]


# --------------------------------------------------------------------------
# numeric helpers
# --------------------------------------------------------------------------

def _mask(t: np.ndarray, lo: float, hi: float) -> np.ndarray:
    return (t >= lo) & (t <= hi)


def _ptp_in(t, x, lo, hi) -> float:
    m = _mask(t, lo, hi)
    if not m.any():
        return 0.0
    seg = x[m]
    if not np.isfinite(seg).all():
        return float("inf")
    return float(np.ptp(seg))


def _bad_values(x: np.ndarray) -> bool:
    return (not np.isfinite(x).all()) or bool((np.abs(x) > 1e6).any())


def scan_log(results_dir: Path, out_stem: str) -> dict:
    """Look for crash/non-convergence markers in the test's .log file."""
    markers = ("not converged", "network not converged", "traceback",
               "error", "failed")
    found = []
    for log in results_dir.rglob("*.log"):
        if out_stem.lower() not in log.name.lower():
            continue
        try:
            text = log.read_text(encoding="utf-8", errors="replace").lower()
        except OSError:
            continue
        found += [m for m in markers if m in text]
    return {"markers": sorted(set(found))}


# --------------------------------------------------------------------------
# per-test evaluations
# --------------------------------------------------------------------------

def _eval_fs(t, chans, criteria, th):
    base = criteria["pu_base"]
    P = _signal(chans, criteria, "P", "FS")
    Q = _signal(chans, criteria, "Q", "FS")
    drift_P = float(np.max(np.abs(P - P[0]))) / base
    drift_Q = float(np.max(np.abs(Q - Q[0]))) / base
    metrics = {"drift_P_pu": round(drift_P, 4), "drift_Q_pu": round(drift_Q, 4)}
    ok = drift_P <= th["drift_P_pu"] and drift_Q <= th["drift_Q_pu"]
    notes = "" if ok else "model does not hold steady output with no disturbance"
    return ok, metrics, notes


def _eval_volt(t, chans, criteria, th, test_name):
    """Voltage step / HVRT tests: worst settled-tail oscillation."""
    base = criteria["pu_base"]
    P = _signal(chans, criteria, "P", test_name)
    Q = _signal(chans, criteria, "Q", test_name)
    pts = read_profile(profile_path(test_name))
    wins = steady_tails(pts, float(t[-1]), criteria["settle_s"],
                        criteria["tail_s"])
    pp_Q = max((_ptp_in(t, Q, lo, hi) for lo, hi in wins), default=0.0) / base
    pp_P = max((_ptp_in(t, P, lo, hi) for lo, hi in wins), default=0.0) / base
    P0 = abs(float(P[0]))
    tail = _mask(t, float(t[-1]) - criteria["settle_s"], float(t[-1]))
    P_end = float(np.mean(P[tail])) if tail.any() else float(P[-1])
    recovered = P_end >= th["p_recovery_frac"] * P0 if P0 > 1e-3 else True

    metrics = {"pp_Q_pu": round(pp_Q, 4), "pp_P_pu": round(pp_P, 4),
               "P_end_frac": round(P_end / P0, 4) if P0 > 1e-3 else None}
    bad = _bad_values(P) or _bad_values(Q)
    ok = (not bad) and pp_Q <= th["pp_Q_pu"] and pp_P <= th["pp_P_pu"] and recovered
    notes = []
    if bad:
        notes.append("non-finite/diverged channel values")
    if pp_Q > th["pp_Q_pu"]:
        notes.append("sustained reactive-power oscillation in steady windows")
    if pp_P > th["pp_P_pu"]:
        notes.append("sustained active-power oscillation in steady windows")
    if not recovered:
        notes.append("active power did not recover after the event")
    return ok, metrics, "; ".join(notes)


def _eval_lvrt(t, chans, criteria, th, test_name):
    """LVRT: momentary-cessation duration plus the VOLT oscillation checks."""
    base = criteria["pu_base"]
    P = _signal(chans, criteria, "P", test_name)
    pts = read_profile(profile_path(test_name))

    fault_times = [tt for tt, v in pts if v < criteria["fault_v_pu"]]
    fault_start = min(fault_times) if fault_times else float(t[0])
    P0 = abs(float(P[0]))
    dt = np.diff(t, prepend=t[0])
    ceased = (np.abs(P) < criteria["cessation_frac"] * max(P0, 1e-6)) & (t >= fault_start)
    cessation_s = float(np.sum(dt[ceased]))

    ok_volt, metrics, notes = _eval_volt(t, chans, criteria, th, test_name)
    metrics["cessation_s"] = round(cessation_s, 2)
    ok = ok_volt and cessation_s <= th["cessation_s"]
    if cessation_s > th["cessation_s"]:
        notes = (notes + "; " if notes else "") + (
            f"momentary cessation {cessation_s:.2f}s exceeds {th['cessation_s']}s "
            "(power does not recover promptly after voltage recovery)")
    return ok, metrics, notes


def _eval_freq(t, chans, criteria, th, test_name):
    """Frequency step tests: oscillation + active-power collapse check."""
    base = criteria["pu_base"]
    P = _signal(chans, criteria, "P", test_name)
    Q = _signal(chans, criteria, "Q", test_name)
    pts = read_profile(profile_path(test_name))
    wins = steady_tails(pts, float(t[-1]), criteria["settle_s"],
                        criteria["tail_s"])
    pp_P = max((_ptp_in(t, P, lo, hi) for lo, hi in wins), default=0.0) / base
    pp_Q = max((_ptp_in(t, Q, lo, hi) for lo, hi in wins), default=0.0) / base
    P0 = abs(float(P[0]))
    tail = _mask(t, float(t[-1]) - criteria["settle_s"], float(t[-1]))
    P_end = float(np.mean(P[tail])) if tail.any() else float(P[-1])
    collapsed = P0 > 1e-3 and P_end < th["collapse_frac"] * P0

    metrics = {"pp_P_pu": round(pp_P, 4), "pp_Q_pu": round(pp_Q, 4),
               "P_end_frac": round(P_end / P0, 4) if P0 > 1e-3 else None}
    bad = _bad_values(P) or _bad_values(Q)
    ok = (not bad) and pp_P <= th["pp_P_pu"] and pp_Q <= th["pp_Q_pu"] and not collapsed
    notes = []
    if bad:
        notes.append("non-finite/diverged channel values")
    if collapsed:
        notes.append("active-power collapse during frequency disturbance")
    if pp_P > th["pp_P_pu"] or pp_Q > th["pp_Q_pu"]:
        notes.append("sustained oscillation in steady windows")
    return ok, metrics, "; ".join(notes)


def _eval_scr2(t, chans, criteria, th, expected_end: float):
    """Weak-grid test: per-SCR-level stability.

    The .out file carries an explicit 'SCR' channel recording the commanded
    level vs time, so segmentation needs no assumptions about switch times.
    A level is judged on the last `tail_s` seconds of its dwell: by then a
    stable response has settled, an unstable one is still oscillating.
    """
    base = criteria["pu_base"]
    V = _signal(chans, criteria, "V", "SCR2")
    Q = _signal(chans, criteria, "Q", "SCR2")
    S = _signal(chans, criteria, "SCR", "SCR2")
    cfg = criteria["scr2"]
    tail = cfg["tail_s"]
    fault_margin = cfg["fault_margin_s"]   # a 4-cycle 3P fault fires at the
    graded = cfg["graded_levels"]          # end of every non-final dwell
    t_end = float(t[-1])

    levels = {}
    detail = {}
    for level in config.SCR2_LEVELS:
        m_level = np.abs(S - level) < 1e-6
        if not m_level.any():
            levels[str(level)] = "unstable"      # sim died before this level
            detail[str(level)] = "level never reached"
            continue
        seg_t = t[m_level]
        seg_start, seg_end = float(seg_t[0]), float(seg_t[-1])
        is_last = level == config.SCR2_LEVELS[-1]
        full_dwell = (seg_end - seg_start) >= config.SCR2_DWELL_S - 0.5
        if not full_dwell:
            levels[str(level)] = "unstable"
            detail[str(level)] = f"simulation ended at t={seg_end:.2f}s mid-dwell"
            continue
        hi = seg_end if is_last else seg_end - fault_margin
        m = _mask(t, hi - tail, hi)
        v_seg, q_seg = V[m], Q[m]
        if v_seg.size == 0 or _bad_values(v_seg) or _bad_values(q_seg):
            levels[str(level)] = "unstable"
            detail[str(level)] = "non-finite/diverged values"
            continue
        pp_v = float(np.ptp(v_seg))
        pp_q = float(np.ptp(q_seg)) / base
        stable = pp_v <= th["pp_V_pu"] and pp_q <= th["pp_Q_pu"]
        levels[str(level)] = "stable" if stable else "unstable"
        detail[str(level)] = f"tail pp_V={pp_v:.3f} pp_Q={pp_q:.3f}"

    metrics = {"levels": levels, "level_detail": detail,
               "graded_levels": graded, "t_end_s": round(t_end, 2)}
    bad_graded = [str(l) for l in graded if levels.get(str(l)) != "stable"]
    ok = not bad_graded
    notes = ("" if ok else
             f"sustained oscillation / instability at graded SCR level(s): "
             f"{', '.join(bad_graded)}")
    ungraded = [str(l) for l in config.SCR2_LEVELS if l not in graded]
    if ungraded:
        notes += (f"{'; ' if notes else ''}level(s) {', '.join(ungraded)} are "
                  "informational only (not stabilizable by the reference solution)")
    if t_end < expected_end - 0.5:
        notes += f"; simulation terminated early at t={t_end:.2f}s"
    return ok, metrics, notes


# --------------------------------------------------------------------------
# suite evaluation
# --------------------------------------------------------------------------

def _expected_t_end(test_name: str, criteria: dict) -> float:
    if test_name == "Test01_FS":
        return 10.0
    if test_name == "Test08_SCR2":
        return criteria["scr2"]["expected_end_s"]
    return max(t for t, _ in read_profile(profile_path(test_name)))


def evaluate_suite(archive_dir: Path, gains: dict, iteration: int,
                   suite_status: str = "completed",
                   elapsed_s: float | None = None,
                   criteria: dict | None = None) -> dict:
    criteria = criteria or load_criteria()
    results_dir = Path(archive_dir) / "RESULTs"
    tests = {}

    for test_name in config.TEST_NAMES:
        th = criteria["tests"][test_name]
        entry = {"status": "error", "metrics": {}, "thresholds": th, "notes": ""}
        try:
            out_file = (_find_out_file(results_dir, test_name, criteria)
                        if results_dir.exists() else None)
            if suite_status != "completed" and out_file is None:
                entry["notes"] = f"suite {suite_status}; no output produced"
                tests[test_name] = entry
                continue
            if out_file is None:
                entry["notes"] = "no .out file produced (test crashed or skipped)"
                log_info = scan_log(results_dir, test_name.split("_", 1)[1])
                if log_info["markers"]:
                    entry["notes"] += f"; log markers: {log_info['markers']}"
                tests[test_name] = entry
                continue

            t_arr, chans = channels.load_channels(out_file)
            t = np.asarray(t_arr, dtype=float)
            expected_end = _expected_t_end(test_name, criteria)

            if test_name == "Test01_FS":
                ok, metrics, notes = _eval_fs(t, chans, criteria, th)
            elif test_name in ("Test02_VOLTDOWN", "Test03_VOLTUP", "Test06_HVRT"):
                ok, metrics, notes = _eval_volt(t, chans, criteria, th, test_name)
            elif test_name == "Test07_LVRT":
                ok, metrics, notes = _eval_lvrt(t, chans, criteria, th, test_name)
            elif test_name in ("Test04_FRQDOWN", "Test05_FRQUP"):
                ok, metrics, notes = _eval_freq(t, chans, criteria, th, test_name)
            elif test_name == "Test08_SCR2":
                ok, metrics, notes = _eval_scr2(t, chans, criteria, th, expected_end)
            else:
                raise EvalError(f"no evaluator for {test_name}")

            # early termination overrides a numeric pass (except SCR2, which
            # already folds termination into its per-level verdicts)
            if test_name != "Test08_SCR2" and float(t[-1]) < expected_end - 0.5:
                ok = False
                notes = (notes + "; " if notes else "") + (
                    f"simulation terminated early at t={float(t[-1]):.2f}s "
                    f"(expected {expected_end:.1f}s)")

            entry.update(status="pass" if ok else "fail",
                         metrics=metrics, notes=notes)
        except (EvalError, Exception) as e:  # noqa: BLE001
            entry["status"] = "error"
            entry["notes"] = f"{type(e).__name__}: {e}"
        tests[test_name] = entry

    n_pass = sum(1 for v in tests.values() if v["status"] == "pass")
    report = {
        "iteration": iteration,
        "gains": gains,
        "suite_status": suite_status,
        "tests": tests,
        "overall_pass": n_pass == len(config.TEST_NAMES),
        "n_pass": n_pass,
        "n_fail": len(config.TEST_NAMES) - n_pass,
    }
    if elapsed_s is not None:
        report["elapsed_s"] = round(elapsed_s, 1)
    return report


def main():
    """CLI: evaluate an archived suite folder.

    python evaluator.py <archive_dir> [<Kqp> <Kqi> <Kvp> <Kvi>] [--iter N]
    """
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("archive_dir")
    ap.add_argument("gains", nargs="*", type=float)
    ap.add_argument("--iter", type=int, default=0)
    args = ap.parse_args()
    gains = (dict(zip(config.GAIN_NAMES, args.gains))
             if len(args.gains) == 4 else {})
    report = evaluate_suite(Path(args.archive_dir), gains, args.iter)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
